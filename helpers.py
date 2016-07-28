import copy
import openpyxl as opyxl

global row_dst 
row_dst = 3

global num_enteries
num_enteries = 30

def get_num_participants():
    workbook = opyxl.load_workbook('Ponds_v_2.xlsx')
    sheet = workbook.get_sheet_by_name('Ponds_v_2.csv')
    num = 0
    i = 0
    while sheet.cell(row = row_dst+i, column = 11).value != None:
        num += 1
        i += 1
    return num

global num_participants 
num_participants = get_num_participants()

def representsInt(s):
    try: 
        int(s)
        return True
    except ValueError:
        return False

def prefix_int(s, k):
	new_s = ""
	for i in xrange(k+1, len(s)):
		if representsInt(s[i:i+1]):
			new_s = new_s + s[i:i+1]
		else :
			return int(new_s)
	return int(new_s)

def get_int(s, i):
	j = 0
	k = 0
	while k < len(s):
		if s[k] == "|":
			j += 1
		if j == i+1:
			return prefix_int(s, k)
		k += 1

def list_from_strings(s):
    if s == "ponds":
        col_dst = 12
    elif s == "queries":
        col_dst = 11
    else:
        assert s == "vall"
        col_dst = 16
    pond_array = []
    workbook = opyxl.load_workbook('Ponds_v_2.xlsx')
    sheet = workbook.get_sheet_by_name('Ponds_v_2.csv')
    result_array = []
    for x in xrange(num_participants):
        new_array = []
        for y in xrange(num_enteries):
            new_array += [None]
        result_array += [copy.deepcopy(new_array)]
    for i in xrange(num_participants):
        pond_array.append((sheet.cell(row = row_dst+i, column = col_dst).value))
    for j in xrange(num_participants*num_enteries):
        pond = get_int(pond_array[j/num_enteries], j%num_enteries)
        result_array[j/num_enteries][j%num_enteries] = pond
    return result_array

def interleave(pond_order1, queries1, vall1):
    assert(len(pond_order1) == len(queries1) and len(queries1) == len(vall1))
    result = copy.deepcopy(pond_order1)
    for i in xrange(len(pond_order1)):
        for j in xrange(len(pond_order1[i])):
            result[i][j] = (pond_order1[i][j], queries1[i][j], vall1[i][j])
    for item in result:
        item.sort()
    return result

def get_data():
	col_dst = 24
	data_array = []
	for x in xrange(num_enteries):
		new_array = []
		for y in xrange(num_participants):
			new_array += [None]
		data_array += [copy.deepcopy(new_array)]
	workbook = opyxl.load_workbook('Ponds_v_2.xlsx')
	sheet = workbook.get_sheet_by_name('Ponds_v_2.csv')
	for i in xrange(num_enteries):
		for j in xrange(num_participants):
			data_array[i][j] = sheet.cell(row = row_dst+j, column = col_dst+i).value
	return data_array

def get_times():
	col_dst = 55
	times_array = []
	for x in xrange(num_enteries):
		new_array = []
		for y in xrange(num_participants):
			new_array += [None]
		times_array += [copy.deepcopy(new_array)]
	workbook = opyxl.load_workbook('Ponds_v_2.xlsx')
	sheet = workbook.get_sheet_by_name('Ponds_v_2.csv')
	for i in xrange(num_enteries):
		for j in xrange(num_participants):
			times_array[i][j] = sheet.cell(row = row_dst+j, column = col_dst + (4*i)).value
	return times_array

def get_catches():
	workbook = opyxl.load_workbook('Ponds_v_2.xlsx')
	sheet = workbook.get_sheet_by_name('Ponds_v_2.csv')
	col_dst = 174
	catch1 = []
	catch2 = []
	for i in xrange(num_participants):
		catch1.append(sheet.cell(row = row_dst+i, column = col_dst).value)
		catch2.append(sheet.cell(row = row_dst+i, column = col_dst+5).value)
	return [catch1, catch2]

def listToString(L, dimension):
	result = "["
	for item in L:
		if dimension == 1:
			result = result+" "+str(item)+","
		elif dimension == 2:
			result = result+" "+listToString(item, 1)+","
	result = result[0:len(result)-1]
	result = result+"]"
	return result

with open("data.py", "w") as f:
	f.write("data_array = "+listToString(get_data(), 2)+"\n")
	f.write("times_array = "+listToString(get_times(), 2)+"\n")
	f.write("catches = "+listToString(get_catches(), 2)+"\n")
	f.write("interleaved = "+listToString(interleave(list_from_strings("ponds"), 
		list_from_strings("queries"), list_from_strings("vall")), 2)+"\n")
	f.write("num_participants = %d" % num_participants+"\n")
	f.write("num_enteries = %d" % num_enteries+"\n\n")
	f.write("class data:\n")
	f.write("    def __init__(self, data_array, times_array, catches_array, "+
		"interleaved_array, num_participants, num_enteries):\n")
	f.write("        self.data_array = data_array\n        self.times = times_array\n"+
		"        self.catches = catches_array\n        self.interleaved = interleaved_array\n"+
		"        self.num_part = num_participants\n        self.num_ent = num_enteries\n\n")
	f.write("info = data(data_array, times_array, catches, interleaved, num_participants, num_enteries)")